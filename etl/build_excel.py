import duckdb
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

conn = duckdb.connect('data/processed/hospital_ops.duckdb')

queries = {
    'KPI Summary': open('sql/analysis_01_kpi_summary.sql').read().strip().rstrip(';'),
    'Top & Bottom Hospitals': open('sql/analysis_02_top_bottom_hospitals.sql').read().strip().rstrip(';'),
    'Anomaly Detection': open('sql/analysis_03_anomaly_detection.sql').read().strip().rstrip(';'),
    'Regional Benchmarking': open('sql/analysis_04_regional_benchmarking.sql').read().strip().rstrip(';'),
    'Readmission Risk': open('sql/analysis_05_readmission_analysis.sql').read().strip().rstrip(';'),
    'Safety Percentile': open('sql/analysis_06_safety_performance.sql').read().strip().rstrip(';'),
}

OUTPUT = 'docs/hospital_ops_analysis.xlsx'

with pd.ExcelWriter(OUTPUT, engine='openpyxl') as writer:
    for sheet_name, query in queries.items():
        df = conn.execute(query).df()
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f'  {sheet_name}: {len(df):,} rows')

# Style the workbook
wb = load_workbook(OUTPUT)

header_font     = Font(bold=True, color='FFFFFF', size=11)
header_fill     = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
alt_fill        = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
center_align    = Alignment(horizontal='center', vertical='center')
thin_border     = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for sheet in wb.sheetnames:
    ws = wb[sheet]

    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Alternate row shading
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = 'A2'

wb.save(OUTPUT)
conn.close()
print(f'\nExcel workbook saved: {OUTPUT}')
